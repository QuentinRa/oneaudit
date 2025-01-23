from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from oneaudit.api.utils.caching import args_api_config, args_parse_api_config, get_cached_result
from oneaudit.modules.leaks.parse import email_formats
from oneaudit.utils.logs import args_verbose_config, args_parse_parse_verbose, get_project_logger
from oneaudit.modules.osint.subdomains.dump import compute_result as dump_subdomains
from oneaudit.modules.osint.hosts.scan import compute_result as port_scan
from oneaudit.modules.socosint.linkedin.scrap import compute_result as find_employees_raw
from oneaudit.modules.socosint.linkedin.export import compute_result as find_employees_export
from oneaudit.modules.socosint.linkedin.parse import compute_result as find_employees_parse
from oneaudit.modules.leaks.parse import compute_result as generate_targets
from oneaudit.modules.leaks.download import compute_result as download_leaks
from oneaudit.modules.leaks.clean import compute_result as clean_leaks
from oneaudit.utils.sheet import create_workbook, workbook_add_sheet_with_table
from os.path import exists as file_exists
from os import makedirs
from json import dump as json_dump


def define_args(parent_parser):
    automate_module = parent_parser.add_parser('automate', help='Automated Module')
    automate_module.add_argument('-d', '--domain', dest='company_domain', help='For example, "example.com".', required=True)
    automate_module.add_argument('-s', '--scope', help='IPs to scan, comma-separated, such as <found>,127.0.0.1/24,127.0.0.1.')
    automate_module.add_argument('-f', '--format', dest='email_format', help='Format used to generate company emails.', choices=email_formats.keys())
    automate_module.add_argument('--alias', dest='domain_aliases', default=[], action='append', help='Alternative domain names that should be investigated.')
    automate_module.add_argument('--restrict', dest='only_from_the_target_domain', action='store_true', help='Only keep emails ending with the provided domain/aliases.')
    automate_module.add_argument('--qleaks', dest='can_download_leaks', action='store_true', help='To avoid download leaks with an incorrect email format, by default, leaks are not downloaded unless you use this switch after confirming the email format.')
    automate_module.add_argument('-o', '--output', dest='output_folder', required=True)
    args_api_config(automate_module)
    args_verbose_config(automate_module)

def run(args):
    args_parse_parse_verbose(args)
    api_keys = args_parse_api_config(args)
    logger = get_project_logger()
    output_folder = f'{args.output_folder}/{args.company_domain}'

    if not file_exists(output_folder):
        makedirs(output_folder, exist_ok=True)

    workbook = create_workbook()
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Subdomains
    subdomain_output_file = f'{output_folder}/subdomains.json'
    args.company_domain = args.company_domain
    args.output_file = subdomain_output_file
    data = dump_subdomains(args, api_keys)
    workbook_add_sheet_with_table(
        workbook=workbook,
        title="Subdomains",
        columns=["Domain", "IP", "ASN", "ASN Name", "ASN Range"],
        rows=[[d.domain_name, d.ip_address, d.asn['asn_id'] if d.asn else "", d.asn['asn_range'] if d.asn else "", d.asn['asn_name'] if d.asn else ""] for d in data['domains']],
        sizes=(50, 25, 10, 20, 15),
        validation_rules=[
            None,
            None,
            None,
            None,
            None,
        ],
        formatting_rules=[
            None,
            [FormulaRule(formula=['ISBLANK(B2)'], fill=bad_fill)],
            [FormulaRule(formula=['ISBLANK(C2)'], fill=bad_fill)],
            [FormulaRule(formula=['ISBLANK(D2)'], fill=bad_fill)],
            [FormulaRule(formula=['ISBLANK(E2)'], fill=bad_fill)],
        ],
        autowrap=False
    )

    # Open Ports
    if args.scope:
        port_scanning_output_file = f'{output_folder}/hosts.json'
        args.target_ips = (args.scope  if "<found>" not in args.scope else args.scope.replace("<found>", ",".join(
            set([d.ip_address for d in data['domains'] if d.ip_address])
        ))).split(',')
        args.output_file = port_scanning_output_file
        args.domains_file = subdomain_output_file
        data = port_scan(args, api_keys)
        workbook_add_sheet_with_table(
            workbook=workbook,
            title="Hosts",
            columns=["IP", "Port", "Domains", "Stack", "Vulnerabilities"],
            rows=[
                [ip_address, port, '\n'.join(host['domains']), '\n'.join(host['stack']), '\n'.join(host['vulns'])]
                for ip_address, host in data['hosts'].items() for port in (host['ports'] if host['ports'] else [None])
            ],
            sizes=(25, 10, 50, 50, 25),
            validation_rules=[
                None,
                None,
                None,
                None,
                None,
            ],
            formatting_rules=[
                None,
                [FormulaRule(formula=['ISBLANK(B2)'], fill=bad_fill)],
                None,
                None,
                None,
            ],
            autowrap=True,
        )
    else:
        logger.warning("Please use --scope to enable port scanning. To scan all IPs found during enumeration, use: -s '<found>'.")

    # Employees
    osint_search_output_file = f'{output_folder}/osint.1.json'
    osint_export_output_file = f'{output_folder}/osint.2.json'
    osint_contacts_output_file = f'{output_folder}/osint.3.json'
    args.company_domain = args.company_domain
    args.company_profile = None
    args.target_profile_list_id = 'auto'
    args.output_file = osint_search_output_file
    find_employees_raw(args, api_keys)

    # RocketReach Export
    profile_list_id = get_cached_result("rocketreach", f"rocketreach_profile_id_{args.company_domain}")
    if profile_list_id:
        args.file_source = 'rocketreach'
        args.profile_list_id = profile_list_id['id']
        args.output_file = osint_export_output_file
        find_employees_export(args, api_keys)

        args.input_file = osint_export_output_file
        args.api_name = 'rocketreach'
        args.filters = ['auto']
        args.output_file = osint_contacts_output_file
        find_employees_parse(args, api_keys)

    targets_output_file = f'{output_folder}/targets.json'
    if args.email_format:
        args.input_files = [osint_search_output_file, osint_contacts_output_file]
        args.domain = args.company_domain
        args.email_format = args.email_format
        args.domain_aliases = args.domain_aliases
        args.only_from_the_target_domain = args.only_from_the_target_domain
        args.output_file = targets_output_file
        targets = generate_targets(args, api_keys)

        workbook_add_sheet_with_table(
            workbook=workbook,
            title="Employees",
            columns=["Login", "Employed", "Verified", "Contacts", "Websites", "Title", "Company"],
            rows=[
                [
                    target.login,
                    target.employed,
                    target.verified,
                    "\n".join(target.emails).strip(),
                    "\n".join(target.extra['links'].values()).strip(),
                    target.extra['current_title'],
                    target.extra['current_company'],
                ]
                for target in targets['credentials']
            ],
            sizes=(50, 15, 15, 50, 100, 50, 25),
            validation_rules=[
                None,
                DataValidation(type="list", formula1='"TRUE,FALSE"', showDropDown=False),
                DataValidation(type="list", formula1='"TRUE,FALSE"', showDropDown=False),
                None,
                None,
                None,
                None,
            ],
            formatting_rules=[
                None,
                [
                    FormulaRule(formula=['B2=TRUE'], fill=good_fill),
                    FormulaRule(formula=['B2=FALSE'], fill=bad_fill),
                ],
                [
                    FormulaRule(formula=['C2=TRUE'], fill=good_fill),
                    FormulaRule(formula=['C2=FALSE'], fill=bad_fill),
                ],
                None,
                None,
                None,
                None,
            ],
            autowrap=True,
        )
    else:
        logger.warning("Please use --format to enable employee email generation and add the 'Targets' sheet.")

    if args.can_download_leaks:
        if not file_exists(targets_output_file):
            logger.warning("No target file, only some API will be able to work...")
            targets_output_file = f"{targets_output_file}.fake"
            with open(targets_output_file, 'w') as output_file:
                json_dump({"credentials": []}, output_file)

        leaks_download_output_file = f'{output_folder}/leaks.0.json'
        args.input_file = targets_output_file
        args.company_domain = args.company_domain
        args.can_use_cache_even_if_disabled = True
        args.run_clean = False
        args.output_file = leaks_download_output_file
        download_leaks(args, api_keys)

        leaks_output_file = f'{output_folder}/leaks.1.json'
        args.input_file = leaks_download_output_file
        args.output_file = leaks_output_file
        leaks = clean_leaks(args, api_keys)

        workbook_add_sheet_with_table(
            workbook=workbook,
            title="Leaks",
            columns=["Login", "Employed", "Verified", "Passwords", "Hashes", "InfoStealers", "Breaches"],
            rows=[
                [
                    leak['login'],
                    leak['employed'],
                    leak['verified'],
                    "\n".join(leak['passwords']).strip(),
                    "\n".join([l["value"] for l in leak['hashes']]+leak['censored_passwords']).strip(),
                    "\n".join([f"{b['computer_name']} ({b['operating_system']}/{b['date_compromised']})" for b in leak['info_stealers']]),
                    "\n".join([f"{b.source} ({b.date})" for b in leak['breaches']]),
                ]
                for leak in leaks['credentials'] if leak['passwords'] or leak['censored_passwords'] or leak['hashes']
            ],
            sizes=(50, 15, 15, 20, 20, 50, 100),
            validation_rules=[
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
            formatting_rules=[
                None,
                [
                    FormulaRule(formula=['B2=TRUE'], fill=good_fill),
                    FormulaRule(formula=['B2=FALSE'], fill=bad_fill),
                ],
                [
                    FormulaRule(formula=['C2=TRUE'], fill=good_fill),
                    FormulaRule(formula=['C2=FALSE'], fill=bad_fill),
                ],
                [FormulaRule(formula=['ISBLANK(B2)'], fill=bad_fill)],
                None,
                None,
                None,
            ],
            autowrap=True,
        )
    else:
        logger.warning("Please use --qleaks to enable leak investigation and add the 'Leaks' sheet.")

    # ws.merge_cells(start_row=1, start_column=1, end_row=len(ports_status), end_column=1)

    # Save
    args.output_file = f'{output_folder}/report.xlsx'
    workbook.save(args.output_file)


