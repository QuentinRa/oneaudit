from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import TableStyleInfo, Table
from openpyxl.utils import get_column_letter

generic_table_style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

def create_workbook():
    """Create an empty workbook with no default sheet"""
    wb = Workbook()
    wb.remove(wb.active)
    return wb

def workbook_add_sheet_with_table(workbook, title, columns, rows, sizes, validation_rules, formatting_rules):
    global generic_table_style
    worksheet = workbook.create_sheet(title=title)

    # Add our data
    worksheet.append(columns)
    for row in rows:
        worksheet.append(row)

    # Apply the style on our data
    table_range = f'{get_column_letter(1)}1:{get_column_letter(len(columns))}{len(rows)+1}'
    table = Table(displayName=f"{title}Table", ref=table_range)
    table.tableStyleInfo = generic_table_style
    worksheet.add_table(table)

    # Apply size restrictions on each column
    for index, size in enumerate(sizes):
        worksheet.column_dimensions[get_column_letter(index + 1)].width = size

    # Add data validation rules
    for index, data_validation in enumerate(validation_rules):
        if not data_validation:
            continue
        worksheet.add_data_validation(data_validation)

        col = get_column_letter(index + 1)
        col_range = f"{col}1:{col}{len(rows)+1}"
        data_validation.add(col_range)

    # Add formatting rules
    for index, col_formatting_rules in enumerate(formatting_rules):
        if not col_formatting_rules:
            continue

        col = get_column_letter(index + 1)
        col_range = f"{col}2:{col}{len(rows)+1}"
        [worksheet.conditional_formatting.add(col_range, formatting_rule) for formatting_rule in col_formatting_rules]
